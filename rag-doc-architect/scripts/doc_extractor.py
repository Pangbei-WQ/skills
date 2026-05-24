import os
import io
import json
import argparse
import pandas as pd
from openpyxl import load_workbook
from PIL import Image

def extract_excel_data(file_path, output_dir):
    """提取 Excel 中的文本数据和图片"""
    result = {
        "file_name": os.path.basename(file_path),
        "sheets": {}
    }
    
    # 提取文本数据 (使用 pandas)
    try:
        xls = pd.ExcelFile(file_path)
        for sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name)
            # 过滤掉全空的 DataFrame
            if not df.dropna(how='all').empty:
                result["sheets"][sheet_name] = {
                    "type": "data",
                    "content": json.loads(df.to_json(orient="records", force_ascii=False))
                }
    except Exception as e:
        print(f"Error reading text from {file_path}: {e}")

    # 提取图片 (使用 openpyxl)
    try:
        wb = load_workbook(file_path, data_only=True)
        img_idx = 1
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if hasattr(ws, '_images') and ws._images:
                sheet_images = []
                for img in ws._images:
                    try:
                        image = Image.open(io.BytesIO(img._data()))
                        img_filename = f"{os.path.splitext(os.path.basename(file_path))[0]}_sheet_{sheet_name}_img_{img_idx}.png"
                        img_path = os.path.join(output_dir, img_filename)
                        image.save(img_path)
                        sheet_images.append(img_path)
                        img_idx += 1
                    except Exception as img_e:
                        print(f"Error saving image in {file_path} sheet {sheet_name}: {img_e}")
                
                if sheet_images:
                    if sheet_name not in result["sheets"]:
                        result["sheets"][sheet_name] = {}
                    result["sheets"][sheet_name]["images"] = sheet_images
    except Exception as e:
         print(f"Error reading images from {file_path}: {e}")
         
    return result

def process_directory(input_dir, output_dir):
    """处理目录下所有的目标文件"""
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    all_results = []
    
    for filename in os.listdir(input_dir):
        file_path = os.path.join(input_dir, filename)
        if not os.path.isfile(file_path):
            continue
            
        # 针对不同类型文件的路由分配 (后续可扩展 word, pdf)
        if filename.endswith(('.xlsx', '.xls')):
            print(f"Processing Excel file: {filename}")
            res = extract_excel_data(file_path, output_dir)
            all_results.append(res)
        else:
            print(f"Skipping unsupported file type: {filename}")
            
    # 输出总的汇总索引文件
    summary_file = os.path.join(output_dir, "batch_extraction_summary.json")
    with open(summary_file, 'w', encoding='utf-8') as f:
        json.dump(all_results, f, ensure_ascii=False, indent=2)
        
    print(f"\nBatch processing complete. Summary saved to {summary_file}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Batch document structure and image extractor.")
    parser.add_argument("input_dir", help="Directory containing documents to process")
    parser.add_argument("--output_dir", default="./extracted_data", help="Directory to save extracted images and summary JSON")
    args = parser.parse_args()
    
    process_directory(args.input_dir, args.output_dir)
