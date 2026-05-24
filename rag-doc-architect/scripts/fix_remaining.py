import os
import shutil
import openpyxl
from openpyxl.utils import range_boundaries
import zipfile
import xml.etree.ElementTree as ET

source_folder = r'C:\Users\michael\Desktop\拆开新增财经部-财经经理-出纳'
tmp_folder = r'c:\Users\michael\.gemini\antigravity\tmp_excel'
output_dir = r"c:\Users\michael\.gemini\antigravity"
target_dir = r"\\192.168.1.6\效率工程部\AI专用知识库\公司业务语境知识库V1\00.待归档文件"

files_to_process = [
    "财务经理财管方向-职位说明书.xlsx", 
    "财务经理财管方向-任职资格.xlsx", 
    "财务经理财管方向-八爪鱼图.xlsx", 
    "财务经理税务方向-八爪鱼图.xlsx"
]

def save_and_copy(md_lines, title):
    md_content = "\n".join(md_lines)
    local_path = os.path.join(output_dir, f"{title}.md")
    with open(local_path, "w", encoding="utf-8") as f:
        f.write(md_content)
    print(f"[Done] Generated locally: {local_path}")
    
    if not os.path.exists(target_dir):
        try: os.makedirs(target_dir, exist_ok=True)
        except Exception: pass
        
    target_path = os.path.join(target_dir, f"{title}.md")
    try:
        shutil.copy2(local_path, target_path)
        print(f"[Copied] Successfully copied to network share: {target_path}")
    except Exception as e:
        print(f"[Error] Failed to write {title}.md to network share. {e}")

# ... (Reuse the same logic for the 3 processors)
def process_qualification(path, filename):
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb[wb.sheetnames[0]]
    
    merged_values = {}
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                merged_values[(r, c)] = top_left_cell_value

    def cell_val(r, c):
        val = merged_values.get((r, c), sheet.cell(row=r, column=c).value)
        if val is None: return ""
        return str(val).strip().replace('\n', ' ')

    data = {}
    for r in range(4, 50):
        zhiji = cell_val(r, 1)
        if not zhiji: continue
        pji = cell_val(r, 2)
        key = f"{zhiji} ({pji})"
        if key not in data:
            data[key] = {"素质": [], "知识": [], "技能": []}
            
        sz_xh = cell_val(r, 4)
        sz_nr = cell_val(r, 5)
        if sz_xh and sz_nr: data[key]["素质"].append(f"{sz_xh}. {sz_nr}")
            
        zs_xh = cell_val(r, 6)
        zs_cd = cell_val(r, 7)
        zs_nr = cell_val(r, 8)
        if zs_xh and zs_nr:
            prefix = f"[{zs_cd}] " if zs_cd else ""
            data[key]["知识"].append(f"{zs_xh}. {prefix}{zs_nr}")
            
        jn_xh = cell_val(r, 9)
        jn_nr = cell_val(r, 10)
        if jn_xh and jn_nr: data[key]["技能"].append(f"{jn_xh}. {jn_nr}")
        
    title = f"2026-财经部-任职资格-{filename.replace('-任职资格.xlsx', '')}"
    lines = [
        "---",
        f'title: "{title}"',
        "type: 职位说明书",
        "tags:",
        "  - 财经与法务",
        "  - 任职资格",
        "rag_optimized: true",
        "---",
        ""
    ]
    for k, content in data.items():
        if not any([content["素质"], content["知识"], content["技能"]]): continue
        lines.append(f"## {k} 任职资格")
        lines.append("")
        if content["素质"]:
            lines.append("### 核心素质")
            for item in content["素质"]: lines.append(f"- {item}")
            lines.append("")
        if content["知识"]:
            lines.append("### 专业知识")
            for item in content["知识"]: lines.append(f"- {item}")
            lines.append("")
        if content["技能"]:
            lines.append("### 业务技能")
            for item in content["技能"]: lines.append(f"- {item}")
            lines.append("")
            
    lines.extend([
        "## 🧠 RAG 智能索引层",
        "",
        "Q: 这份任职资格文档主要解决什么问题？",
        f"A: 本文档提供了{filename.replace('-任职资格.xlsx', '')}的任职资格模型，详细规定了各职级所需的“核心素质”、“专业知识”与“业务技能”。",
        ""
    ])
    save_and_copy(lines, title)

def process_jd(path, filename):
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb[wb.sheetnames[0]]
    
    title = f"2026-财经部-JD-{filename.replace('.xlsx', '')}"
    lines = [
        "---",
        f'title: "{title}"',
        "type: 职位说明书",
        "tags:",
        "  - 财经与法务",
        "rag_optimized: true",
        "---",
        ""
    ]
    
    current_h2 = "基本信息"
    info_dict = {}
    
    for r in range(1, 100):
        row_vals = [str(sheet.cell(row=r, column=c).value or "").strip() for c in range(1, 10)]
        row_vals = [v for v in row_vals if v]
        
        if not row_vals: continue
        
        if len(row_vals) == 1:
            val = row_vals[0]
            if len(val) < 20 and ("职位" in val or "基本" in val or "职责" in val or "要求" in val or "工作" in val or "目的" in val):
                if info_dict:
                    lines.append(f"## {current_h2}")
                    lines.append("")
                    for k, v in info_dict.items():
                        lines.append(f"- **{k}**: {v}")
                    lines.append("")
                    info_dict = {}
                current_h2 = val
            else:
                info_dict[f"内容条目"] = val
        else:
            if len(row_vals) >= 2:
                for i in range(0, len(row_vals)-1, 2):
                    info_dict[row_vals[i]] = row_vals[i+1]
                if len(row_vals) % 2 != 0:
                    info_dict[f"内容_{r}"] = row_vals[-1]
    
    if info_dict:
        lines.append(f"## {current_h2}")
        lines.append("")
        for k, v in info_dict.items():
            lines.append(f"- **{k}**: {v}")
        lines.append("")
        
    lines.extend([
        "## 🧠 RAG 智能索引层",
        "",
        "Q: 这份文档主要解决什么问题？",
        f"A: 本文档提供了{filename.replace('.xlsx', '')}的具体JD要求与职责明细。",
        ""
    ])
    
    save_and_copy(lines, title)

def process_octopus(path, filename):
    title = f"2026-财经部-脑图-{filename.replace('.xlsx', '')}"
    z = zipfile.ZipFile(path)
    drawing_files = [f for f in z.namelist() if f.startswith('xl/drawings/drawing')]
    
    texts = []
    namespaces = {'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'}
    for df in drawing_files:
        xml_content = z.read(df)
        root = ET.fromstring(xml_content)
        texts.extend([t.text for t in root.findall('.//a:t', namespaces) if t.text and str(t.text).strip()])
    
    lines = [
        "---",
        f'title: "{title}"',
        "type: 知识结构图",
        "tags:",
        "  - 财经与法务",
        "  - 八爪鱼图",
        "rag_optimized: true",
        "---",
        "",
        "## 结构化脑图内容 (GSA提炼节点)",
        ""
    ]
    
    if texts:
        for t in texts:
            lines.append(f"- {t}")
    else:
        lines.append("- (未提取到图形文本内容，可能包含非常规 SmartArt 结构)")
        
    lines.append("")
    lines.extend([
        "## 🧠 RAG 智能索引层",
        "",
        "Q: 本文档提供了什么核心信息？",
        f"A: 本文档是{filename.replace('.xlsx', '')}的八爪鱼图，展平提取了其中的视觉层级结构文本。",
        ""
    ])
    
    save_and_copy(lines, title)


os.makedirs(tmp_folder, exist_ok=True)
for f in files_to_process:
    src = os.path.join(source_folder, f)
    dst = os.path.join(tmp_folder, f)
    print(f"\n---> Trying {f}")
    if os.path.exists(src):
        try:
            shutil.copy2(src, dst)
            if '任职资格' in f:
                process_qualification(dst, f)
            elif '职位说明书' in f:
                process_jd(dst, f)
            elif '八爪鱼图' in f:
                process_octopus(dst, f)
        except Exception as e:
            print(f"Error processing {f}: {e}")
    else:
        print(f"Source file not found: {f}")
