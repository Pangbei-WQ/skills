import openpyxl
import os
import shutil
from openpyxl.utils import range_boundaries

file_path = r'C:\Users\michael\Desktop\拆开新增财经部-财经经理-出纳\财务经理税务方向-任职资格.xlsx'
try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
except Exception:
    import sys; sys.exit(1)

sheet = wb[wb.sheetnames[0]]

merged_values = {}
for merged_range in sheet.merged_cells.ranges:
    min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
    top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            merged_values[(row, col)] = top_left_cell_value

def cell_val(r, c):
    val = merged_values.get((r, c), sheet.cell(row=r, column=c).value)
    if val is None: return ""
    return str(val).strip().replace('\n', ' ')

data = {} # { "一级|P1|知": { "素质": [], "知识": [], "技能": [] } }

for r in range(4, 50):
    zhiji = cell_val(r, 1)
    if not zhiji: continue
    
    pji = cell_val(r, 2)
    key = f"{zhiji} ({pji})"
    
    if key not in data:
        data[key] = {"素质": [], "知识": [], "技能": []}
    
    sz_xh = cell_val(r, 4)
    sz_nr = cell_val(r, 5)
    if sz_xh and sz_nr: data[key]["素质"].append(f"{sz_xh}. {sz_nr}")
        
    zs_xh = cell_val(r, 6)
    zs_cd = cell_val(r, 7)
    zs_nr = cell_val(r, 8)
    if zs_xh and zs_nr:
        prefix = f"[{zs_cd}] " if zs_cd else ""
        data[key]["知识"].append(f"{zs_xh}. {prefix}{zs_nr}")
        
    jn_xh = cell_val(r, 9)
    jn_nr = cell_val(r, 10)
    if jn_xh and jn_nr: data[key]["技能"].append(f"{jn_xh}. {jn_nr}")

md_lines = []
md_lines.append("---")
title = "2026-财经部-任职资格-财务经理税务方向"
md_lines.append(f"title: \"{title}\"")
md_lines.append("type: 职位说明书")
md_lines.append("tags:")
md_lines.append("  - 财经与法务")
md_lines.append("  - 职位考评")
md_lines.append("  - 任职资格")
md_lines.append("rag_optimized: true")
md_lines.append("---")
md_lines.append("")

for level_key, content in data.items():
    md_lines.append(f"## {level_key} 任职资格")
    md_lines.append("")
    
    if content["素质"]:
        md_lines.append(f"### 核心素质")
        for item in content["素质"]:
            md_lines.append(f"- {item}")
        md_lines.append("")
        
    if content["知识"]:
        md_lines.append(f"### 专业知识")
        for item in content["知识"]:
            md_lines.append(f"- {item}")
        md_lines.append("")
        
    if content["技能"]:
        md_lines.append(f"### 业务技能")
        for item in content["技能"]:
            md_lines.append(f"- {item}")
        md_lines.append("")

# Append RAG section
md_lines.append("## 🧠 RAG 智能索引层")
md_lines.append("")
md_lines.append("Q: 这份文档主要解决什么问题？")
md_lines.append("A: 本文档提供了财务经理(税务方向)的任职资格模型，详细规定了从一级(P1)到四级(P4)各职级所需的“核心素质”、“专业知识”与“业务技能”。")
md_lines.append("")
md_lines.append("Q: 一级(P1)职级的核心素质有哪些要求？")
md_lines.append("A: 一级(P1)职级的核心素质要求基础的学习能力、较强的主动性（无需催促，主动准备）、团队合作（接受别人意见）、初步的思维能力、诚信担当，以及专注超越（做事有始有终）。")
md_lines.append("")
md_lines.append("Q: 不同层级在《税务局预警事项处理》上面有什么不同要求？")
md_lines.append("A: P1和P2层级主要强调“了解”税务处理及预警事项、协助排查或者独立沟通处理简单疑点；P3层级需要“熟悉”政府部门核查处理方法，攥写自查报告并指导下级；P4层级要求达到“精通”，能够维护工商税务政府部门合作关系并处理公司对外涉税事项。")
md_lines.append("")
md_lines.append("Q: 作为四级(P4)在团队建设上有什么要求？")
md_lines.append("A: P4级别的财务经理需要掌握团队维护和建设，负责组员招聘一面、带教、谈话，并能够根据各组员特点分配工作并维护组内和谐合作关系。")
md_lines.append("")
md_lines.append("Q: 各职级对吉客云ERP的掌握要求是什么？")
md_lines.append("A: 职级由浅入深，P1要求“了解”，P2需“掌握”熟练下载和筛查异常；高级别则重点放在流程优化与深层次核算模型的设计上面。")
md_lines.append("")

output_md = "\n".join(md_lines)
local_out_path = f"c:\\Users\\michael\\.gemini\\antigravity\\{title}.md"
with open(local_out_path, "w", encoding="utf-8") as f:
    f.write(output_md)
print(f"Generated Markdown at {local_out_path}")

target_dir = r"\\192.168.1.6\效率工程部\AI专用知识库\公司业务语境知识库V1\00.待归档文件"
target_path = os.path.join(target_dir, f"{title}.md")

try:
    if not os.path.exists(target_dir):
        os.makedirs(target_dir, exist_ok=True)
    shutil.copy2(local_out_path, target_path)
    print(f"Successfully copied to {target_path}")
except Exception as e:
    print(f"Failed to write to network share {target_dir}. Error: {e}")
