import openpyxl
import sys
from openpyxl.utils import range_boundaries

file_path = r'C:\Users\michael\Desktop\拆开新增财经部-财经经理-出纳\财务经理税务方向-任职资格.xlsx'

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
except Exception as e:
    print("Error loading:", e)
    sys.exit(1)

for sheet_name in wb.sheetnames:
    print(f"=== Sheet: {sheet_name} ===")
    sheet = wb[sheet_name]
    
    merged_values = {}
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged_values[(row, col)] = top_left_cell_value
                
    for row_idx, row in enumerate(sheet.iter_rows(), start=1):
        row_data = []
        for col_idx, cell in enumerate(row, start=1):
            val = merged_values.get((row_idx, col_idx), cell.value)
            val = str(val).replace('\n', ' | ') if val is not None else ""
            row_data.append(val)
        
        if any(v.strip() for v in row_data):
            print(f"Row {row_idx}: " + " || ".join(row_data))
    
    print("\n")
